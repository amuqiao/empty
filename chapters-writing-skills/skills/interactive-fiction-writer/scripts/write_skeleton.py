#!/usr/bin/env python3
"""
write_skeleton.py — Read results JSON and write speaker/content/emotion back into xlsx.

Standalone CLI. No dependencies on other scripts in this directory.

Column mapping (0-indexed):
  0=Text(row type)  1=Choice letter  2=Object marker
  3=Narration/Speaker  4=Content  5=Emotion
"""

import argparse
import json
import sys
from typing import List

import pandas as pd


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ALLOWED_EMOTIONS = {
    "angry", "smile", "laughing", "flirty", "neutral",
    "sad", "shocked", "confused", "eyesclosed", "shy",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_tab(path: str, tab: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=tab, header=None)


def get_cell(row: pd.Series, col: int) -> str:
    v = row.iloc[col] if col < len(row) else None
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def apply_results(df: pd.DataFrame, results: List[dict]) -> None:
    for col in [3, 4, 5]:
        if col in df.columns and df[col].dtype != object:
            df[col] = df[col].astype(object)
    result_map = {r["row_index"]: r for r in results}
    for idx in df.index:
        if idx in result_map:
            r = result_map[idx]
            # speaker is only written back if the result provides it (branch slots)
            if r.get("speaker"):
                df.at[idx, 3] = r["speaker"]
            df.at[idx, 4] = r.get("content", "")
            df.at[idx, 5] = r.get("emotion", "")


def validate_output(df: pd.DataFrame) -> List[str]:
    warnings = []
    consec_narration = 0
    seen: set = set()

    for _, row in df.iterrows():
        col_a = get_cell(row, 0)
        col_d = get_cell(row, 3)
        col_e = get_cell(row, 4)
        col_f = get_cell(row, 5)

        if col_f and col_f not in ALLOWED_EMOTIONS:
            warnings.append(f"Row {row.name}: invalid emotion '{col_f}'")

        if col_d == "Narration":
            consec_narration += 1
            if consec_narration >= 4:
                warnings.append(f"Row {row.name}: 4+ consecutive Narration rows")
        else:
            consec_narration = 0

        if col_e and col_d != "Choice":
            if col_e in seen:
                warnings.append(f"Row {row.name}: repeated line '{col_e[:60]}'")
            seen.add(col_e)

        # Check branch slot rows have speaker
        if col_a == "Branch slot" and not col_d:
            warnings.append(f"Row {row.name}: branch slot missing speaker")

        # Check for unfilled GENERATE placeholders
        if col_e == "GENERATE" or col_d == "GENERATE":
            warnings.append(f"Row {row.name}: unfilled GENERATE placeholder")

    return warnings


def write_xlsx(skeleton_df: pd.DataFrame, result_indices: set,
               input_path: str, output_path: str, skeleton_tab: str) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(input_path)

    # Remove all tabs except the skeleton tab
    for name in wb.sheetnames:
        if name != skeleton_tab:
            del wb[name]

    ws = wb[skeleton_tab]

    # Insert a header row at the top (shifts existing data down by 1)
    ws.insert_rows(1)
    headers = ["Text", "Choice letter", "Object marker",
               "Narration", "Content", "Emotion", "Generated"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    # Write generated content into the (now shifted) data rows
    for df_idx, row in skeleton_df.iterrows():
        xlsx_row = int(df_idx) + 2  # +2: 1-indexed + inserted header
        d_val = row.iloc[3]
        e_val = row.iloc[4]
        f_val = row.iloc[5]
        ws.cell(row=xlsx_row, column=4,
                value=str(d_val) if pd.notna(d_val) and d_val != "" else None)
        ws.cell(row=xlsx_row, column=5,
                value=str(e_val) if pd.notna(e_val) and e_val != "" else None)
        ws.cell(row=xlsx_row, column=6,
                value=str(f_val) if pd.notna(f_val) and f_val != "" else None)
        # Column G: mark model-generated rows
        if df_idx in result_indices:
            ws.cell(row=xlsx_row, column=7, value=True)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Read results JSON and write speaker/content/emotion back into xlsx."
    )
    parser.add_argument("--input", required=True, help="Path to input xlsx (source skeleton)")
    parser.add_argument("--results", required=True, help="Path to results JSON file")
    parser.add_argument("--output", required=True, help="Output xlsx path (never overwrites source)")
    parser.add_argument("--skeleton-tab", default="Skeleton", help="Tab name for skeleton")
    parser.add_argument("--dry-run", action="store_true", help="Validate without writing output file")
    args = parser.parse_args()

    with open(args.results, "r", encoding="utf-8") as f:
        results = json.load(f)

    skeleton_df = load_tab(args.input, args.skeleton_tab)

    if args.dry_run:
        print(f"[write] DRY RUN — would write {len(results)} results to {args.output}", file=sys.stderr)
        return

    result_indices = {r["row_index"] for r in results}
    apply_results(skeleton_df, results)
    warnings = validate_output(skeleton_df)
    if warnings:
        for w in warnings:
            print(f"[WARN] {w}", file=sys.stderr)
    write_xlsx(skeleton_df, result_indices, args.input, args.output, args.skeleton_tab)
    print(f"[write] Wrote {len(results)} results to {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
